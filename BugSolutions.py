import openpyxl
from openpyxl.utils import get_column_letter
import numpy as np

import re

import requests
from bs4 import BeautifulSoup as soup
from random import randrange
from datetime import timedelta


# Get cell coordinates for a specific value
class CustomFix:

    ##---------------------------Third Party Starts----------------------------------------------------##
    ############################################
    # import openpyxl
    # from openpyxl.utils import get_column_letter
    # import numpy as np
    ###########################################
    # The CustomFix class is imported from:
    # (https://stackoverflow.com/questions/71603830/openpyxl-module-python-how-can-i-get-the-coordinates-of-a-cell-when-i-know-th)
    def __init__(self, _my_file_name="Banking User Info.xlsx"):
        # assert _my_file_name.split('.')[-1] == 'xlsx', 'Input file is not xlsx'
        self.my_filename = _my_file_name
        self.my_base_wb = openpyxl.load_workbook(self.my_filename, read_only=False)
        # following line will set the last worksheet in the workbook as active
        self.my_base_active_ws = self.my_base_wb.active

    # Method to get values for specific row in a given worksheet
    # Argument to this method is: - Row number of values to be fetched
    def get_specific_row_val_as_list_in_active_ws(self, _val_row_num):
        for _col in self.my_base_active_ws.iter_cols(min_col=1, max_col=1):
            # Iterate once for the specific row number in active worksheet
            for _row in self.my_base_active_ws.iter_rows(min_row=_val_row_num, max_row=_val_row_num):
                # Return a list of values
                return [_cell.value for _cell in _row]

    # Method to get cell coordinate by a search value
    # Argument to this method is:- search string
    # Assumption cell value is unique
    # _search_val == the value currently in the cell  you are looking for
    # first and last name are used only to confirm the user being found
    def get_cell_coordinate_by_UNIQUE_value(self, _search_val):
        # List comprehension to get the row index based on the search value
        _row_processor = [_row_idx for _row_idx, _main_rec in enumerate(self.my_base_active_ws.values, start=1) if
                          _search_val in _main_rec]
        # return type is a list, hence following line to assign it to variable and manage the data type later
        _row_idx = _row_processor[-1]
        # Get the value of the entire row and fetch the column index
        _col_processor = [_col_idx for _col_idx, _val in
                          enumerate(self.get_specific_row_val_as_list_in_active_ws(int(_row_idx)), start=1) if
                          _val == _search_val]
        # return type is a list, hence following line to assign it to variable and manage the data type later
        _col_idx = _col_processor[-1]
        # get the column letter
        _col_letter = get_column_letter(int(_col_idx))
        # string concatenation to join column letter and row index
        _cell_address = _col_letter + str(_row_idx)
        return _cell_address

    def get_cell_coordinate_by_TWO_values(self, _search_val, _first_name, _last_name):
        # List comprehension to get the row index based on the search value
        _row_processor = [_row_idx for _row_idx, _main_rec in enumerate(self.my_base_active_ws.values, start=1) if
                          np.logical_and.reduce(
                              (_first_name in _main_rec, _last_name in _main_rec, _search_val in _main_rec))]
        # return type is a list, hence following line to assign it to variable and manage the data type later
        _row_idx = _row_processor[-1]
        # Get the value of the entire row and fetch the column index
        _col_processor = [_col_idx for _col_idx, _val in
                          enumerate(self.get_specific_row_val_as_list_in_active_ws(int(_row_idx)), start=1) if
                          _val == _search_val]
        # return type is a list, hence following line to assign it to variable and manage the data type later
        _col_idx = _col_processor[-1]
        # get the column letter
        _col_letter = get_column_letter(int(_col_idx))
        # string concatenation to join column letter and row index
        _cell_address = _col_letter + str(_row_idx)
        return _cell_address

    ##--------------------------Third Party ends----------------------------------------------------------##
    ################################
    # Input address info or changes
    ################################
    def get_street_num(self):
        _streetNO = int(input("Enter house number: \n"))
        if len(str(_streetNO)) in range(4):
            return _streetNO
        else:
            print("Not a valid street number...!")
            return CustomFix.get_street_num()

    def get_street_name(self):
        try:
            _streetName = input("Enter street name: \n")
            return _streetName
        except ValueError:
            print("Stree name can not be numeric...!")
            return CustomFix().get_street_name()

    def get_city_name(self):
        try:
            cityName = input("Enter city name: \n")
            return cityName
        except ValueError:
            print("City name can not be numeric...!")
            return CustomFix().get_city_name()

    def get_state_name(self):
        try:
            state = input("Enter state:\n")
            return state
        except ValueError:
            print("State name can not be numeric...!")
            return CustomFix().get_state_name()

    def get_zipcode(self):
        zip = int(input("Enter zipcode: \n"))
        if len(str(zip)) == 5:
            return zip
        else:
            print("Not a valid zipcode number...!")
            return CustomFix().get_zipcode()

    ################################
    # Change a cell value
    ################################
    # This will update a cell value from excel and save the file again
    # it neeed to use the openpyexl library
    def change_excel_cell_value(self, _cell_address, _newValue):
        _exFile = openpyxl.load_workbook("Banking User Info.xlsx")
        sheet = _exFile.active
        # sheet = _exFile.get_sheet_by_name("Sheet1")
        sheet[_cell_address] = _newValue
        _exFile.save("Banking User Info.xlsx")

    ################################
    # Remove an account
    ################################
    def remove_account(self, row_num):
        _exFile = openpyxl.load_workbook("Banking User Info.xlsx")
        sheet = _exFile.active
        sheet.delete_rows(row_num)
        _exFile.save("Banking User Info.xlsx")

    # -----------The following is to fill the Excel file with random values only-------------------
    # --------------------------This is NOT (Third Party) ---------------------
    #######################################
    # import re for the following
    # perimeter validation
    ######################################
    # regex match client names
    def re_match_name(self, _name):
        if re.search("([A-Z][a-z]+.)+", _name):
            return _name
        else:
            print("Not a valid name...")
            return CustomFix().re_match_name(input("Please try again! \n"))

    # regex validate client emails
    def validate_email(self, _user_email):
        _valid_email = re.findall("[a-zA-Z0-9]+@[a-z]+\.[a-z]{2,3}", _user_email)
        if _valid_email:
            return _valid_email[0]
        else:
            print("Not a valid email address...")
            return CustomFix().validate_email(input("Try entering a different email! \n"))

    def _validate_value_in_excel_file(self, _search_value, email=False):
        # if the input is an email (validate it)
        if email == True:
            try:
                _valid_email = CustomFix().validate_email(_search_value)  # email validator
                value_cordinates = CustomFix().get_cell_coordinate_by_UNIQUE_value(_search_value)
                return value_cordinates
            except IndexError:
                print("Your input can be found in our system!")
                return CustomFix()._validate_value_in_excel_file(input("Please try again:\n"), True)

        else:  # if search value is not an email (just go find its coordinates)
            try:
                value_cordinates = CustomFix().get_cell_coordinate_by_UNIQUE_value(_search_value)
                return value_cordinates
            except IndexError:
                return "Your input can be found in our system!"

    #########################################
    # import requests
    # from bs4 import BeautifulSoup as soup
    # from random import randrange
    # from datetime import timedelta
    ########################################
    def random_date(self, start, end):
        """
        This function will return a random datetime between two datetime
        objects.
        """
        delta = end - start
        int_delta = (delta.days * 24 * 60 * 60) + delta.seconds
        random_second = randrange(int_delta)
        return start + timedelta(seconds=random_second)

    def auto_generate_zipcode(self):
        zipcodes = [45160, 49209, 91452, 38971, 25081, 64963, 79852, 55320, 98852, 74790,
                    49422, 71599, 25227, 24182, 75541, 32628, 74553, 97380, 95144, 63425]
        # for ele in range(20):
        #     zipcodes.append(random.randint(10000, 99999))
        return zipcodes

    def auto_fill_client_city(self):
        florida_cities = ['Alachua', 'Altamonte Springs', 'Anna Maria', 'Apalachicola', 'Apopka', 'Atlantic Beach',
                          'Auburndale', 'Aventura', 'Avon Park', 'Bal Harbour', 'Bartow', 'Bay Harbor Islands',
                          'Boca Raton', 'Bonita Springs', 'Boynton Beach', 'Bradenton', 'Brooksville', 'Cape Canaveral',
                          'Cape Coral', 'Casselberry', 'Celebration', 'Chipley', 'Cinco Bayou', 'Clearwater',
                          'Clermont',
                          'Clewiston', 'Cocoa', 'Cocoa Beach', 'Coconut Creek', 'Coral Gables', 'Coral Springs',
                          'Crystal River', 'Dania Beach', 'Davie', 'Daytona Beach', 'Deerfield Beach',
                          'DeFuniak Springs',
                          'DeLand', 'Delray Beach', 'Deltona', 'Destin', 'Dunedin', 'Eagle Lake', 'Edgewater',
                          'Edgewood',
                          'Eustis', 'Fort Lauderdale', 'Fort Meade', 'Fort Myers', 'Fort Myers Beach', 'Fort Pierce',
                          'Fort Walton Beach', 'Fruitland Park', 'Gainesville', 'Greenacres', 'Green Cove Springs',
                          'Gulf Breeze', 'Gulfport', 'Haines City']
        # city_url = "https://www.stateofflorida.com/cities/"
        # data = requests.get(city_url)
        #
        # html_data = soup(data.text, "html.parser")
        # _cityData = html_data.find("div", {"class": "col-md-4"}).findAll("li")
        # for ele in _cityData:
        #     florida_cities.append(ele.find("a").getText())
        return florida_cities

    def auto_fill_client_street(self):
        streets = ['12th Street', '14th Street', '16th Avenue', '16th Street', '18th Avenue', '20th Street',
                   '24th Street', '25th Street', '28th Avenue', '28th Street', '29th Street', '2nd Avenue',
                   '2nd Avenue',
                   '2nd Boulevard', '3 Mile Drive', '30th Street', '31st Street', '32nd Avenue', '32nd Street',
                   '33rd Street', '35th Street', '3rd Avenue', '3rd St', '4 Tops Drive', '51st Street', '52nd Street',
                   '8 Mile Road West', 'Aaron Street', 'Abandoned Deimler-chrysler G O', 'Abbott Street', 'Abington',
                   'Abington Avenue', 'Acacia Avenue', 'Ackley Street', 'Adair Street', 'Addison Street',
                   'Adelaide Street',
                   'Adeline Street', 'Ailey Court', 'Akron Avenue', 'Alameda Street', 'Alaska Avenue', 'Alaska Street',
                   'Albany Avenue', 'Albany Street', 'Albion Avenue', 'Albion Street', 'Alcoy Avenue', 'Alcoy Street',
                   'Alden Street', 'Alderton Street', 'Alexandrine Street West', 'Alfred Street', 'Alger Street',
                   'Algonac Avenue', 'Algonac Street', 'Algonquin Park Drive', 'Algonquin Street', 'All Saints Street',
                   'Allen Place', 'Allendale Street', 'Allonby Street', 'Alma Avenue', 'Alma Street', 'Almont Avenue',
                   'Almont Street', 'Alpena Avenue', 'Alpine Street', 'Alter Road', 'Alwar Street', 'Ambassador Bridge',
                   'Ambassador Bridge Street', 'American Avenue', 'American Street', 'American Way Street',
                   'Amity Street',
                   'Amrad Street', 'Amsterdam Street', 'Anatole Street', 'Anderdon Avenue', 'Anderdon Street',
                   'Anderson Street', 'Anglin Street', 'Annabelle Street', 'Annchester Road', 'Annin Street',
                   'Annland Street', 'Annott Avenue', 'Annott Street', 'Annsbury Avenue', 'Anson Street',
                   'Anstell Avenue',
                   'Anthon Street', 'Anthony Wayne Drive', 'Antietam Avenue', 'Antoinette Street', 'Antwerp Avenue',
                   'Anvil Avenue', 'Anvil Street', 'Appleton Street', 'Appoline Avenue', 'Appoline Street',
                   'Arcadia Street',
                   'Archdale Street', 'Archer Street', 'Arden Park Boulevard', 'Ardmore Street', 'Aretha Avenue',
                   'Argus Avenue', 'Argyle Crescent', 'Armada Street', 'Armour Street', 'Army Street', 'Arndt Street',
                   'Arnold Avenue', 'Artesian Street', 'Asa Avenue', 'Asbury Park', 'Ashland Street', 'Ashton Avenue',
                   'Ashton Road', 'Astor Street', 'Athens Avenue', 'Atkinson Street', 'Atlanta Street', 'Atwater Drive',
                   'Atwater Street', 'Auburn Street', 'Audrey Street', 'Audubon Road', 'Aurora Street', 'Austin Street',
                   'Averhill Court', 'Avery Street', 'Avery Terrace', 'Avis Street', 'Avon Avenue', 'Avon Road',
                   'Bacon Street', 'Badger Street', 'Bagley Avenue', 'Baldwin Street', 'Balfour Road', 'Balmoral Drive',
                   'Bangor Street', 'Bank Street', 'Banneker Court', 'Barbara Street', 'Barham Avenue', 'Barham Street',
                   'Barker Avenue', 'Barker Street', 'Barlow Avenue', 'Barlow Street', 'Barlum Avenue', 'Barlum Street',
                   'Barnes Street', 'Barr Street', 'Barrett Avenue', 'Barrett Street', 'Barron Street', 'Basil Street',
                   'Bassett Street', 'Bates Street', 'Baubee Avenue', 'Bauman Street', 'Baylis Street',
                   'Beaconsfield Street',
                   'Beals Street', 'Beaman Street', 'Beard Street', 'Beaubien Street', 'Beaverland Street',
                   'Bedford Road',
                   'Bedford Street', 'Beechdale Avenue', 'Beechdale Street', 'Beechton Street', 'Beechwood Street',
                   'Begole Street', 'Beland Street', 'Belden Street', 'Belfast Avenue', 'Belfast Street',
                   'Belle Street',
                   'Belleterre Avenue', 'Belleterre Street', 'Bellevue Street', 'Belton', 'Belvidere Street',
                   'Benham Street',
                   'Beniteau Street', 'Benson Street', 'Benson Street', 'Bentler Court', 'Bentler Street',
                   'Benton Street',
                   'Berden Street', 'Berg Road', 'Berkeley Road', 'Berkshire Road', 'Berry Street', 'Bessemore Street',
                   'Bethune Avenue West', 'Bethune Street West', 'Beverly Court', 'Bewick Street', 'Biltmore Street',
                   'Binder Avenue', 'Binder Street', 'Birchcrest Drive', 'Birwood Avenue', 'Birwood Street',
                   'Bishop Street',
                   'Bivouac Street', 'Blackmoor Street', 'Blackstone Court', 'Blackstone Street', 'Blaine Street',
                   'Blair Avenue', 'Bliss Avenue', 'Bliss Street', 'Bloom Street', 'Blowers Street', 'Bluehill',
                   'Bluehill Street', 'Blythe Street', 'Boleyn Street', 'Bonita Street', 'Bordeau Street',
                   'Borman Avenue',
                   'Borman Street', 'Bortle Avenue', 'Bostwick Street', 'Bosworth Court', 'Boulder Street',
                   'Bourke Street',
                   'Boxwood Street', 'Brace Street', 'Braden Avenue', 'Braden Street', 'Bradford Avenue',
                   'Bradley Street',
                   'Braile Street', 'Brainard Avenue', 'Brainard Street', 'Bramell Street', 'Bramford Street',
                   'Brandon Street', 'Breckenridge Avenue', 'Breckenridge Street', 'Bremen Street',
                   'Brennan Pool Drive',
                   'Brennan Street', 'Bretton', 'Bretton Drive', 'Brewery Park Boulevard', 'Briarcliff Road',
                   'Brimson Avenue', 'Brimson Street', 'Bringard Drive', 'Brinker Avenue', 'Brinker Street',
                   'Brinket Avenue',
                   'Bristow Street', 'Britain Avenue', 'Britain Street', 'Broadstreet Avenue', 'Broadway Street',
                   'Brock Avenue', 'Brockton Street', 'Bromley Avenue', 'Brooklyn Street', 'Brooks Street',
                   'Brown Place',
                   'Bruce St', 'Bruckner Street', 'Brunswick Street', 'Bryanston Crescent Street', 'Bryant Street',
                   'Bryden Street', 'Buchanan Avenue', 'Buckingham Avenue', 'Buelow Court', 'Buhl Street',
                   'Buhr Street',
                   'Bulwer Street', 'Burchill Court', 'Burdeno Street', 'Burgess Court', 'Burgess Street',
                   'Burlage Place',
                   'Burlingame Street', 'Burlington Drive', 'Burnette Avenue', 'Burnette Street', 'Burns Drive',
                   'Burnside Street', 'Burrell Place', 'Burt Court', 'Burt Road', 'Burwell Avenue', 'Burwell Street',
                   'Bushey Street', 'Byron Street', 'C L Franklin Boulevard', 'Cabacier Street', 'Cabot Street',
                   'Cadet Street', 'Cadieux Road', 'Cadillac Assembly Plant', 'Cadillac Boulevard', 'Cadillac Square',
                   'Caely Street', 'Cahalan Street', 'Cairney Street', 'Caldwell Street', 'Calvert Street',
                   'Cambridge Avenue', 'Camden Avenue', 'Camille Street', 'Camley Road', 'Camley Street',
                   'Campau Farms Circle', 'Canal Street', 'Canonbury Avenue', 'Canonbury Street', 'Canterbury Road',
                   'Canton Street', 'Canyon Avenue', 'Canyon Street', 'Capitol Street', 'Carbon Street',
                   'Carbondale Avenue',
                   'Carbondale Street', 'Cardoni Avenue', 'Cardoni Street', 'Carleton Street', 'Carlin Street',
                   'Carlisle Drive', 'Carlisle Street', 'Carol Street', 'Carrie Avenue', 'Carrie Street',
                   'Carson Street',
                   'Carten Street', 'Carter Street', 'Cartridge Street', 'Cary Street', 'Cascade Avenue',
                   'Cascade Street',
                   'Casgrain Street', 'Casino Avenue', 'Casino Way', 'Casper Street', 'Cass @ Plaza Dr', 'Cass Avenue',
                   'Castle Avenue', 'Castleton Street', 'Cathedral Street', 'Cavalry Street', 'Cecil Avenue',
                   'Cecil Street',
                   'Cedargrove Avenue', 'Cedargrove Street', 'Cedarhurst Place', 'Celestine Avenue', 'Celia Street',
                   'Central Street', 'Centre Street', 'Chadwick Street', 'Chalfonte Street', 'Chalmers Street',
                   'Chandler Park Drive', 'Chandler Street', 'Chapel Street', 'Chapin Street', 'Charest Street',
                   'Charlemagne Avenue', 'Charleston Street', 'Charlevoix', 'Charlevoix', 'Charlevoix Crane Cutoff',
                   'Charlevoix Street', 'Charlotte Street', 'Chateaufort Place', 'Chatfield Street', 'Chatham Street',
                   'Chatsworth Road', 'Chatsworth Street', 'Chelsea Avenue', 'Chene Court', 'Chene Street',
                   'Chenlot Street',
                   'Cherboneau Place', 'Cherrylawn Avenue', 'Cherrylawn Avenue', 'Cherrylawn Street', 'Chester Street',
                   'Chesterfield Road', 'Cheyenne Street', 'Childrens Way', 'Chipman Street', 'Chippewa Avenue',
                   'Chippewa Street', 'Chope Place', 'Chopin Street', 'Christiancy Street', 'Christy Avenue',
                   'Christy Street', 'Chrysler Drive', 'Chrysler Expressway', 'Chrysler Freeway', 'Cicotte Street',
                   'Civic Center Drive', 'Civic Place', 'Clairepoint Street', 'Clairepoint Woods Drive',
                   'Clairmount Street',
                   'Clairpointe Avenue', 'Clairpointe Street', 'Clairpointe Woods Drive', 'Clairview Street',
                   'Clarion Street', 'Clarita Avenue', 'Clark Court', 'Clarkdale Street', 'Clayburn Street',
                   'Clayton Street', 'Clements Street', 'Cliff Street', 'Clifford Street', 'Clifton Street',
                   'Clinton River Shore Avenue', 'Clinton River Shore Street', 'Clippert Avenue', 'Clough Street',
                   'Cloverdale Street', 'Cloverlawn Avenue', 'Cloverlawn Street', 'Coastal Drive', 'Cobb Place',
                   'Cochrane Street', 'Codding Street', 'Cody Street', 'Coe Avenue', 'Colfax Avenue', 'Colfax Street',
                   'Collingham Drive', 'Collingwood Avenue', 'Collingwood Street', 'Colton Street', 'Commercial Street',
                   'Commonwealth Street', 'Compass Avenue', 'Compass Street', 'Conant Street', 'Concord Avenue',
                   'Concord Street', 'Conger Street', 'Congress Street', 'Conley Avenue', 'Conley Street',
                   'Conner Lane',
                   'Conner Street', 'Conrad Avenue', 'Conrad Street', 'Conservatory Drive', 'Constance Street',
                   'Continental Street', 'Conway Street', 'Cooley Street', 'Cooper Street', 'Cope Street', 'Copland',
                   'Copland Street', 'Coplin Avenue', 'Coplin Street', 'Coram Avenue', 'Coram Street', 'Corbett Street',
                   'Cordell Street', 'Cordova Street', 'Cornwall Avenue', 'Cornwall Street', 'Cortland Street',
                   'Courville Street', 'Coventry Street', 'Covert Street', 'Coyle Avenue', 'Coyle Street',
                   'Craft Street',
                   'Crane Street', 'Cranshaw Avenue', 'Cranshaw Street', 'Crawford Street', 'Cresswell Street',
                   'Crocuslawn Avenue', 'Crocuslawn Street', 'Crusade Street', 'Cruse Street', 'Cumberland Way',
                   'Curt Street', 'Curtis Avenue', 'Curtis Street', 'Cushing Street', 'Custer Avenue', 'Custer Street',
                   'Cutler Street', 'Cymbal Street', 'D Ann Street', 'Dacosta Street', 'Dailey Avenue', 'Dailey Court',
                   'Dailey Street', 'Dakota Avenue West', 'Dalrymple Street', 'Dalzelle Street', 'Daniels Street',
                   'Darcy Street', 'Davenport Street', 'David Avenue', 'Davis Place', 'Davison Crossover',
                   'Dawes Street',
                   'Dayton Avenue', 'Dayton Street', 'Dean Avenue', 'Dearing Street', 'Debuel Street', 'Decatur Street',
                   'Defer Place', 'Dehner Street', 'Delaware Street', 'Delmar Street', 'Deming Street',
                   'Denmark Street',
                   'Dennis Street', 'Dennison Street', 'Denver Street', 'Dequindre Street', 'Desmond Street',
                   'Desner Avenue', 'Desoto Street', 'Devereaux Avenue', 'Devereaux Street', 'Devine Avenue',
                   'Devine Street', 'Devon Street', 'Devonshire Road', 'Dexter Avenue', 'Dey Street',
                   'Dickerson Street',
                   'Dill Place', 'Distel Street', 'Dix Street', 'Dobel Street', 'Dolphin Street', 'Dolson Street',
                   'Domine Street', 'Donald Place', 'Dorothy Street', 'Dover Avenue', 'Downing Street', 'Doyle Street',
                   'Dragoon Street', 'Dresden Avenue', 'Dresden Street', 'Drew Court', 'Drexel Street',
                   'Drifton Avenue',
                   'Driggs Street', 'Du Charme Place', 'Duane Street', 'Dubay Avenue', 'Dubay Street', 'Dubois Street',
                   'Ducharme Place', 'Duchess Avenue', 'Duchess Street', 'Duffield Street', 'Dumfries Road',
                   'Dumfries Street', 'Dunedin Avenue', 'Dunedin Street', 'Duprey Street', 'Durham Street',
                   'Dwight Street',
                   'Dwyer Avenue', 'Dwyer Street', 'Eagle Street', 'Earle Street', 'East 7 Mile Road',
                   'East 7 Mile Road',
                   'East Adams Street', 'East Alexandrine Street', 'East Baltimore Avenue', 'East Bethune Avenue',
                   'East Borman Avenue', 'East Borman Drive', 'East Boston Boulevard', 'East Brentwood Street',
                   'East Canfield Street', 'East Columbia Street', 'East Congress Street', 'East Crescent Lane',
                   'East Dakota Avenue', 'East Davison Street', 'East Edsel Ford Freeway', 'East Ferry Avenue',
                   'East Fisher Freeway', 'East Fort Street', 'East Golden Gate', 'East Goldengate',
                   'East Grand Boulevard',
                   'East Grand River Avenue', 'East Greendale Street', 'East Grixdale Avenue', 'East Hancock Street',
                   'East Harbortown Drive', 'East Hildale Avenue', 'East Hollywood Avenue', 'East Hollywood Street',
                   'East Jefferson Avenue', 'East Jefferson Avenue', 'East Kirby Avenue', 'East Kirby Street',
                   'East Kirby Street', 'East La Salle Gardens Street', 'East Lantz Avenue', 'East Lantz Street',
                   'East Larned Street', 'East Margaret Street', 'East Mcnichols Road', 'East Milwaukee Street',
                   'East Montana Street', 'East Montcalm Street', 'East Morrow Circle', 'East Nevada Street',
                   'East Palmer Avenue', 'East Parkhurst Place', 'East Philadelphia Avenue', 'East Philadelphia Street',
                   'East Picnic Way', 'East Remington Avenue', 'East Robinwood Street', 'East Savannah Street',
                   'East State Fair', 'East State Fair Street', 'East Street', 'East Vernor Highway',
                   'East Warren Avenue',
                   'East Willis Avenue', 'East Willis Street', 'East Yemans Avenue', 'Eastburn Avenue',
                   'Eastburn Street',
                   'Eastern Place', 'Eastlawn Street', 'Eastwood Avenue', 'Eastwood Street', 'Eaton Avenue',
                   'Edgeton Avenue',
                   'Edgeton Street', 'Edgevale Street', 'Edgewood Avenue', 'Edinborough Road', 'Edison Street',
                   'Edlie Circle', 'Edlie Street', 'Edmonton Street', 'Edmore Drive', 'Edmund Place',
                   'Edsel Ford Freeway',
                   'Edsel Ford Freeway Service Road', 'Edsel Ford Freeway West', 'Edsel Ford Fwy I- 94 Service',
                   'Edward Avenue', 'Edward Street', 'Edwin Drive South', 'Elba Place', 'Eldon Avenue', 'Eldred Street',
                   'Eldridge Street', 'Elgin Avenue', 'Elgin Street', 'Elijah Mccoy Drive', 'Eliot', 'Eliot Street',
                   'Eliza Howell Park', 'Ellen Avenue', 'Ellery Place', 'Ellery Street', 'Ellis Avenue', 'Ellis Street',
                   'Elmdale Street', 'Elmer Avenue', 'Elmer Street', 'Elmhurst Street', 'Elmira Street', 'Elmo Street',
                   'Elmwood Street', 'Elsa Street', 'Elsmere Street', 'Emery Street', 'Emily Avenue', 'Emmons Street',
                   'Endicott Avenue', 'Engle Street', 'Engleside Drive', 'Engleside Street', 'Englewood Street',
                   'Epworth Avenue', 'Epworth Street', 'Erbie Street', 'Erie Street', 'Erskine Street', 'Erwin Avenue',
                   'Esper Street', 'Essex Avenue', 'Estates Drive', 'Ethel Street', 'Eugene Avenue', 'Eureka Street',
                   'Evanston Street', 'Evergreen Avenue', 'Evergreen Road', 'Evergreen Srv Road', 'Everts Avenue',
                   'Everts Street', 'Ewald Circle', 'Ewers Street', 'Exeter Street', 'Fairbanks Street',
                   'Faircrest Avenue',
                   'Faircrest Street', 'Fairfield Street', 'Fairmount Drive', 'Fairport Street', 'Fairview Avenue',
                   'Fairview Street', 'Falcon Street', 'Fargo Avenue', 'Fargo Street', 'Farmbrook Avenue',
                   'Farmbrook Street', 'Farnsworth Avenue', 'Farnsworth Street', 'Farr Street', 'Faust Avenue',
                   'Federal Street', 'Felch Street', 'Fenelon Avenue', 'Fenelon Street', 'Fenkell Street',
                   'Fenmore Avenue',
                   'Fenmore Street', 'Fenton', 'Fenwick Street', 'Ferdinand Street', 'Ferguson Street',
                   'Fernwood Avenue',
                   'Ferris Street', 'Ferry Avenue East', 'Ferry Mall - Wayne State University', 'Ferry Park Avenue',
                   'Ferry Park St', 'Ferry Park Street', 'Field Street', 'Fielding Street', 'Filer Avenue',
                   'Filer Street',
                   'Findlay Avenue', 'Findlay Street', 'Firwood Avenue', 'Firwood Street', 'Fischer Avenue',
                   'Fischer Street',
                   'Fisher Freeway', 'Fisher Frwy Serv Drive', 'Fiske Drive', 'Fitzpatrick Court', 'Fitzpatrick Street',
                   'Fleet Street', 'Fleming Street', 'Florida Street', 'Flower Court', 'Floyd Avenue', 'Floyd Street',
                   'Foley Street', 'Ford Place - Wayne State University', 'Ford Street', 'Fordale Street',
                   'Fordham Avenue',
                   'Fordham Street', 'Fordson Street', 'Fordyce Street', 'Forest Lawn Cemetery', 'Forestlawn Street',
                   'Forrer Street', 'Foster Street', 'Fountain Way', 'Fournier Street', 'Francis Avenue',
                   'Franfort Court',
                   'Frank Street', 'Frankfort Avenue', 'Frankfort Court', 'Frankfort Street', 'Fraser Place',
                   'Frederick Avenue', 'Frederick Douglass Avenue', 'Fredro Street', 'Freeland Street', 'Freer Avenue',
                   'Freer Street', 'Fremont Place', 'French Road', 'Freud Street', 'Frisbee Street', 'Frontenac Street',
                   'Fullerton Street', 'Fulton Street', 'Gable Avenue', 'Gable Street', 'Gainsborough Road',
                   'Gale Street',
                   'Gallagher Street', 'Galster Street', 'Gar Street', 'Gardendale Street', 'Garland Street',
                   'Garnet Avenue', 'Garnet Street', 'Gartner Street', 'Garvin Street', 'Gates Street',
                   'Gateshead Street',
                   'Gavel Street', 'Gaylord Avenue', 'Gaynor Court', 'Genessee Avenue', 'Genoa Street',
                   'Georgeland Street',
                   'Georgia Street', 'Gerisch', 'Gethsemane Cemetary', 'Gibson Street', 'Giese Street',
                   'Gietzen Street',
                   'Gilbert Avenue', 'Gilbo Avenue', 'Gilchrist Street', 'Gillespie Court', 'Gillett Street',
                   'Gilroy Street', 'Girardin Avenue', 'Girardin Street', 'Gitre Avenue', 'Gitre Street',
                   'Gladstone Avenue',
                   'Gladstone Street', 'Gladys Court', 'Gladys Street', 'Glastonbury Avenue', 'Glastonbury Road',
                   'Gleason Street', 'Glenco Avenue', 'Glenco Street', 'Glendale Street', 'Glenfield',
                   'Glenfield Avenue',
                   'Glenhurst Street', 'Glinnan Street', 'Gloucester Drive', 'Glynn Court', 'Goethe Street',
                   'Goldner Street',
                   'Goldsmith Street', 'Goodwin Street', 'Gordon', 'Gore Avenue', 'Goulburn Avenue', 'Goulburn Street',
                   'Govin Street', 'Grace Street', 'Grand River Avenue', 'Grandmont Avenue', 'Grandmont Road',
                   'Grandview Avenue', 'Grandview Street', 'Grandville Avenue', 'Grandy Street', 'Granger Street',
                   'Gratiot Avenue', 'Gratiot Court', 'Graves Avenue', 'Graves Street', 'Gravier Street', 'Gray Street',
                   'Grayfield Street', 'Grayton Street', 'Greater Grace Temple Boulevard', 'Greeley Street',
                   'Green Place',
                   'Greendale Street East', 'Greendale Street West', 'Greenfield Road', 'Greenlawn Avenue',
                   'Greenlawn Street', 'Greenlodge Street', 'Greensboro Avenue', 'Greensboro Street',
                   'Greenview Avenue',
                   'Greenview Road', 'Greenway Avenue', 'Greenway Street', 'Gregorie Street', 'Greiner Street',
                   'Greusel Street', 'Greydale Avenue', 'Greydale Court', 'Greydale Street', 'Griggs Avenue',
                   'Griggs Street', 'Grinnell Avenue', 'Griswold Street', 'Grixdale Avenue East',
                   'Grixdale Avenue West',
                   'Groesbeck Highway', 'Grotto Court', 'Grove Avenue', 'Grover Avenue', 'Gruebner Avenue',
                   'Gruebner Street', 'Guilford Street', 'Gullen Mall - Wayne State University', 'Gunston',
                   'Gunston Avenue',
                   'Gunston Street', 'Guoin Street', 'Guthrie Street', 'Hackett Street', 'Hafeli Street', 'Hague',
                   'Hale Street', 'Hall Avenue', 'Halleck Avenue', 'Halleck Street', 'Halley Street', 'Hamburg Avenue',
                   'Hamburg Street', 'Hamilton Road', 'Hamlet Street', 'Hammond Street', 'Hampshire Street',
                   'Hancock Street East', 'Hanson Street', 'Harbaugh', 'Harbor Island Street', 'Harbortown Drive West',
                   'Hardyke Street', 'Harlow Avenue', 'Harlow Street', 'Harmon Street', 'Harned Street', 'Harnor Court',
                   'Harper I- 94', 'Harrell Street', 'Harrington Street', 'Harry Street', 'Hart Plaza',
                   'Hartford Street',
                   'Hartwell Avenue', 'Hartwell Street', 'Hartwick Street', 'Harvard Road', 'Hasse Avenue',
                   'Hasse Street',
                   'Hastings Street', 'Hathon Street', 'Havana Street', 'Haverhill Street', 'Hazelridge Avenue',
                   'Hazelridge Street', 'Hazlett Avenue', 'Hazlett Street', 'Healy Avenue', 'Healy Street',
                   'Heck Place',
                   'Hecla Street', 'Hedge Street', 'Heidelberg Street', 'Heintz Street', 'Hemlock Street',
                   'Henderson Avenue',
                   'Henderson Street', 'Hendricks Street', 'Hendrie Avenue', 'Hendrie Street']

        # streets_url = "https://geographic.org/streetview/usa/mi/detroit.html"
        # data = requests.get(streets_url)
        # html_data = soup(data.text, "html.parser")
        # _streetData = html_data.body.find("div", {"class": "listmain"}).findAll("li")
        # for ele in range(1000):
        #     streets.append(_streetData[ele].find("a").getText())
        return streets

    def auto_fill_client_names(self):
        names = ['Olivia', 'Emma', 'Charlotte', 'Amelia', 'Ava', 'Sophia', 'Isabella', 'Mia', 'Evelyn', 'Harper',
                 'Luna', 'Camila', 'Gianna', 'Elizabeth', 'Eleanor', 'Ella', 'Abigail', 'Sofia', 'Avery', 'Scarlett',
                 'Emily', 'Aria', 'Penelope', 'Chloe', 'Layla', 'Mila', 'Nora', 'Hazel', 'Madison', 'Ellie', 'Lily',
                 'Nova', 'Isla', 'Grace', 'Violet', 'Aurora', 'Riley', 'Zoey', 'Willow', 'Emilia', 'Stella', 'Zoe',
                 'Victoria', 'Hannah', 'Addison', 'Leah', 'Lucy', 'Eliana', 'Ivy', 'Everly', 'Lillian', 'Paisley',
                 'Elena', 'Naomi', 'Maya', 'Natalie', 'Kinsley', 'Delilah', 'Claire', 'Audrey', 'Aaliyah', 'Ruby',
                 'Brooklyn', 'Alice', 'Aubrey', 'Autumn', 'Leilani', 'Savannah', 'Valentina', 'Kennedy', 'Madelyn',
                 'Josephine', 'Bella', 'Skylar', 'Genesis', 'Sophie', 'Hailey', 'Sadie', 'Natalia', 'Quinn', 'Caroline',
                 'Allison', 'Gabriella', 'Anna', 'Serenity', 'Nevaeh', 'Cora', 'Ariana', 'Emery', 'Lydia', 'Jade',
                 'Sarah', 'Eva', 'Adeline', 'Madeline', 'Piper', 'Rylee', 'Athena', 'Peyton', 'Everleigh', 'Vivian',
                 'Clara', 'Raelynn', 'Liliana', 'Samantha', 'Maria', 'Iris', 'Ayla', 'Eloise', 'Lyla', 'Eliza',
                 'Hadley', 'Melody', 'Julia', 'Parker', 'Rose', 'Isabelle', 'Brielle', 'Adalynn', 'Arya', 'Eden',
                 'Remi', 'Mackenzie', 'Maeve', 'Margaret', 'Reagan', 'Charlie', 'Alaia', 'Melanie', 'Josie', 'Elliana',
                 'Cecilia', 'Mary', 'Daisy', 'Alina', 'Lucia', 'Ximena', 'Juniper', 'Kaylee', 'Magnolia', 'Summer',
                 'Adalyn', 'Sloane', 'Amara', 'Arianna', 'Isabel', 'Reese', 'Emersyn', 'Sienna', 'Kehlani', 'River',
                 'Freya', 'Valerie', 'Blakely', 'Genevieve', 'Esther', 'Valeria', 'Katherine', 'Kylie', 'Norah',
                 'Amaya', 'Bailey', 'Ember', 'Ryleigh', 'Georgia', 'Catalina', 'Emerson', 'Alexandra', 'Faith',
                 'Jasmine', 'Ariella', 'Ashley', 'Andrea', 'Millie', 'June', 'Khloe', 'Callie', 'Juliette',
                 'Sage', 'Ada', 'Anastasia', 'Olive', 'Alani', 'Brianna', 'Rosalie', 'Molly', 'Brynlee', 'Amy', 'Ruth',
                 'Aubree', 'Gemma', 'Taylor', 'Oakley', 'Margot', 'Arabella', 'Sara', 'Journee', 'Harmony', 'Blake',
                 'Alaina', 'Aspen', 'Noelle', 'Selena', 'Oaklynn', 'Morgan', 'Londyn', 'Zuri', 'Aliyah', 'Jordyn',
                 'Juliana', 'Finley', 'Presley', 'Zara', 'Leila', 'Marley', 'Sawyer', 'Amira', 'Lilly', 'London',
                 'Kimberly', 'Elsie', 'Ariel', 'Lila', 'Alana', 'Diana', 'Kamila', 'Nyla', 'Vera', 'Hope', 'Annie',
                 'Kaia', 'Myla', 'Alyssa', 'Angela', 'Ana', 'Lennon', 'Evangeline', 'Harlow', 'Rachel', 'Gracie',
                 'Rowan', 'Laila', 'Elise', 'Sutton', 'Lilah', 'Adelyn', 'Phoebe', 'Octavia', 'Sydney', 'Mariana',
                 'Wren', 'Lainey', 'Vanessa', 'Teagan', 'Kayla', 'Malia', 'Elaina', 'Saylor', 'Brooke', 'Lola',
                 'Miriam', 'Alayna', 'Adelaide', 'Daniela', 'Jane', 'Payton', 'Journey', 'Lilith', 'Delaney',
                 'Dakota', 'Mya', 'Charlee', 'Alivia', 'Annabelle', 'Kailani', 'Lucille', 'Trinity', 'Gia', 'Tatum',
                 'Raegan', 'Camille', 'Kaylani', 'Kali', 'Stevie', 'Maggie', 'Haven', 'Tessa', 'Daphne', 'Adaline',
                 'Hayden', 'Joanna', 'Jocelyn', 'Lena', 'Evie', 'Juliet', 'Fiona', 'Cataleya', 'Angelina', 'Leia',
                 'Paige', 'Julianna', 'Milani', 'Talia', 'Rebecca', 'Kendall', 'Harley', 'Lia', 'Phoenix', 'Dahlia',
                 'Logan', 'Camilla', 'Thea', 'Jayla', 'Brooklynn', 'Blair', 'Vivienne', 'Hallie', 'Madilyn', 'Mckenna',
                 'Evelynn', 'Ophelia', 'Celeste', 'Alayah', 'Winter', 'Catherine', 'Collins', 'Nina', 'Briella',
                 'Palmer', 'Noa', 'Mckenzie', 'Kiara', 'Amari', 'Adriana', 'Gracelynn', 'Lauren', 'Cali', 'Kalani',
                 'Aniyah', 'Nicole', 'Alexis', 'Mariah', 'Gabriela', 'Wynter', 'Amina', 'Ariyah', 'Adelynn',
                 'Remington', 'Reign', 'Alaya', 'Dream', 'Alexandria', 'Willa', 'Avianna', 'Makayla', 'Gracelyn',
                 'Elle', 'Amiyah', 'Arielle', 'Elianna', 'Giselle', 'Brynn', 'Ainsley', 'Aitana', 'Charli', 'Demi',
                 'Makenna', 'Rosemary', 'Danna', 'Izabella', 'Lilliana', 'Melissa', 'Samara', 'Lana', 'Mabel',
                 'Everlee', 'Fatima', 'Leighton', 'Esme', 'Raelyn', 'Madeleine', 'Nayeli', 'Camryn', 'Kira', 'Annalise',
                 'Selah', 'Serena', 'Royalty', 'Rylie', 'Celine', 'Laura', 'Brinley', 'Frances', 'Michelle', 'Heidi',
                 'Rory', 'Sabrina', 'Destiny', 'Gwendolyn', 'Alessandra', 'Poppy', 'Amora', 'Nylah', 'Luciana',
                 'Maisie',
                 'Miracle', 'Joy', 'Liana', 'Raven', 'Shiloh', 'Allie', 'Daleyza', 'Kate', 'Lyric', 'Alicia', 'Lexi',
                 'Addilyn', 'Anaya', 'Malani', 'Paislee', 'Elisa', 'Kayleigh', 'Azalea', 'Francesca', 'Jordan',
                 'Regina', 'Viviana', 'Aylin', 'Skye', 'Daniella', 'Makenzie', 'Veronica', 'Legacy', 'Maia', 'Ariah',
                 'Alessia', 'Carmen', 'Astrid', 'Maren', 'Helen', 'Felicity', 'Alexa', 'Danielle', 'Lorelei', 'Paris',
                 'Adelina', 'Bianca', 'Gabrielle', 'Jazlyn', 'Scarlet', 'Bristol', 'Navy', 'Esmeralda', 'Colette',
                 'Stephanie', 'Jolene', 'Marlee', 'Sarai', 'Hattie', 'Nadia', 'Rosie', 'Kamryn', 'Kenzie', 'Alora',
                 'Holly', 'Matilda', 'Sylvia', 'Cameron', 'Armani', 'Emelia', 'Keira', 'Braelynn', 'Jacqueline',
                 'Alison', 'Amanda', 'Cassidy', 'Emory', 'Ari', 'Haisley', 'Jimena', 'Jessica', 'Elaine', 'Dorothy',
                 'Mira', 'Eve', 'Oaklee', 'Averie', 'Charleigh', 'Lyra', 'Madelynn', 'Angel', 'Edith', 'Jennifer',
                 'Raya', 'Ryan', 'Heaven', 'Kyla', 'Wrenley', 'Meadow', 'Carter', 'Kora', 'Saige', 'Kinley',
                 'Maci', 'Mae', 'Salem', 'Aisha', 'Adley', 'Carolina', 'Sierra', 'Alma', 'Helena', 'Bonnie', 'Mylah',
                 'Briar', 'Aurelia', 'Leona', 'Macie', 'Maddison', 'April', 'Aviana', 'Lorelai', 'Alondra', 'Kennedi',
                 'Monroe', 'Emely', 'Maliyah', 'Ailani', 'Madilynn', 'Renata', 'Katie', 'Zariah', 'Imani', 'Amber',
                 'Analia', 'Ariya', 'Anya', 'Emberly', 'Emmy', 'Mara', 'Maryam', 'Dior', 'Mckinley', 'Virginia',
                 'Amalia', 'Mallory', 'Opal', 'Shelby', 'Clementine', 'Remy', 'Xiomara', 'Elliott', 'Elora', 'Katalina',
                 'Antonella', 'Skyler', 'Hanna', 'Kaliyah', 'Alanna', 'Haley', 'Itzel', 'Cecelia', 'Jayleen', 'Kensley',
                 'Beatrice', 'Journi', 'Dylan', 'Ivory', 'Yaretzi', 'Meredith', 'Sasha', 'Gloria', 'Oaklyn', 'Sloan',
                 'Abby', 'Davina', 'Lylah', 'Erin', 'Reyna', 'Kaitlyn', 'Michaela', 'Nia', 'Fernanda', 'Jaliyah',
                 'Jenna', 'Sylvie', 'Miranda', 'Anne', 'Mina', 'Myra', 'Aleena', 'Alia', 'Frankie', 'Ellis', 'Kathryn',
                 'Nalani', 'Nola', 'Jemma', 'Lennox', 'Marie', 'Angelica', 'Cassandra', 'Calliope', 'Adrianna',
                 'Ivanna',
                 'Zelda', 'Faye', 'Karsyn', 'Oakleigh', 'Dayana', 'Amirah', 'Megan', 'Siena', 'Reina', 'Rhea',
                 'Julieta',
                 'Malaysia', 'Henley', 'Liberty', 'Leslie', 'Alejandra', 'Kelsey', 'Charley', 'Capri', 'Priscilla',
                 'Zariyah', 'Savanna', 'Emerie', 'Christina', 'Skyla', 'Macy', 'Mariam', 'Melina', 'Chelsea', 'Dallas',
                 'Laurel', 'Briana', 'Holland', 'Lilian', 'Amaia', 'Blaire', 'Margo', 'Louise', 'Rosalia', 'Aleah',
                 'Bethany', 'Flora', 'Kylee', 'Kendra', 'Sunny', 'Laney', 'Tiana', 'Chaya', 'Ellianna', 'Milan',
                 'Aliana', 'Estella', 'Julie', 'Yara', 'Rosa', 'Cheyenne', 'Emmie', 'Carly', 'Janelle', 'Kyra', 'Naya',
                 'Malaya', 'Sevyn', 'Lina', 'Mikayla', 'Jayda', 'Leyla', 'Eileen', 'Irene', 'Karina', 'Aileen', 'Aliza',
                 'Kataleya', 'Kori', 'Indie', 'Lara', 'Romina', 'Jada', 'Kimber', 'Amani', 'Liv', 'Treasure', 'Louisa',
                 'Marleigh', 'Winnie', 'Kassidy', 'Noah', 'Monica', 'Keilani', 'Zahra', 'Zaylee', 'Hadassah', 'Jamie',
                 'Allyson', 'Anahi', 'Maxine', 'Karla', 'Khaleesi', 'Johanna', 'Penny', 'Hayley', 'Marilyn', 'Della',
                 'Freyja', 'Jazmin', 'Kenna', 'Ashlyn', 'Florence', 'Ezra', 'Melany', 'Murphy', 'Sky', 'Marina',
                 'Noemi',
                 'Coraline', 'Selene', 'Bridget', 'Alaiya', 'Angie', 'Fallon', 'Thalia', 'Rayna', 'Martha', 'Halle',
                 'Estrella', 'Joelle', 'Kinslee', 'Roselyn', 'Theodora', 'Jolie', 'Dani', 'Elodie', 'Halo', 'Nala',
                 'Promise', 'Justice', 'Nellie', 'Novah', 'Estelle', 'Jenesis', 'Miley', 'Hadlee', 'Janiyah', 'Waverly',
                 'Braelyn', 'Pearl', 'Aila', 'Katelyn', 'Sariyah', 'Azariah', 'Bexley', 'Giana', 'Lea', 'Cadence',
                 'Mavis', 'Ila', 'Rivka', 'Jovie', 'Yareli', 'Bellamy', 'Kamiyah', 'Kara', 'Baylee', 'Jianna', 'Kai',
                 'Alena', 'Novalee', 'Elliot', 'Livia', 'Ashlynn', 'Denver', 'Emmalyn', 'Persephone', 'Marceline',
                 'Jazmine', 'Kiana', 'Mikaela', 'Aliya', 'Galilea', 'Harlee', 'Jaylah', 'Lillie', 'Mercy', 'Ensley',
                 'Bria', 'Kallie', 'Celia', 'Berkley', 'Ramona', 'Jaylani', 'Jessie', 'Aubrie', 'Madisyn', 'Paulina',
                 'Averi', 'Aya', 'Chana', 'Milana', 'Cleo', 'Iyla', 'Cynthia', 'Hana', 'Lacey', 'Andi', 'Giuliana',
                 'Milena', 'Leilany', 'Saoirse', 'Adele', 'Drew', 'Bailee', 'Hunter', 'Rayne', 'Anais', 'Kamari',
                 'Paula', 'Rosalee', 'Teresa', 'Zora', 'Avah', 'Belen', 'Greta', 'Layne', 'Scout', 'Zaniyah', 'Amelie',
                 'Dulce', 'Chanel', 'Clare', 'Rebekah', 'Giovanna', 'Ellison', 'Isabela', 'Kaydence', 'Rosalyn',
                 'Royal', 'Alianna', 'August', 'Nyra', 'Vienna', 'Amoura', 'Anika', 'Harmoni', 'Kelly', 'Linda',
                 'Aubriella', 'Kairi', 'Ryann', 'Avayah', 'Gwen', 'Whitley', 'Noor', 'Khalani', 'Marianna', 'Addyson',
                 'Annika', 'Karter', 'Vada', 'Tiffany', 'Artemis', 'Clover', 'Laylah', 'Paisleigh', 'Elyse', 'Kaisley',
                 'Veda', 'Zendaya', 'Simone', 'Alexia', 'Alisson', 'Angelique', 'Ocean', 'Elia', 'Lilianna', 'Maleah',
                 'Avalynn', 'Marisol', 'Goldie', 'Malayah', 'Emmeline', 'Paloma', 'Raina', 'Brynleigh', 'Chandler',
                 'Valery', 'Adalee', 'Tinsley', 'Violeta', 'Baylor', 'Lauryn', 'Marlowe', 'Birdie', 'Jaycee', 'Lexie',
                 'Loretta', 'Lilyana', 'Princess', 'Shay', 'Hadleigh', 'Natasha', 'Indigo', 'Zaria', 'Addisyn',
                 'Deborah', 'Leanna', 'Barbara', 'Kimora', 'Emerald', 'Raquel', 'Julissa', 'Robin', 'Austyn', 'Dalia',
                 'Nyomi', 'Ellen', 'Kynlee', 'Salma', 'Luella', 'Zayla', 'Addilynn', 'Giavanna', 'Samira', 'Amaris',
                 'Madalyn', 'Scarlette', 'Stormi', 'Etta', 'Ayleen', 'Brittany', 'Brylee', 'Araceli', 'Egypt', 'Iliana',
                 'Paityn', 'Zainab', 'Billie', 'Haylee', 'India', 'Kaiya', 'Nancy', 'Clarissa', 'Mazikeen', 'Taytum',
                 'Aubrielle', 'Rylan', 'Ainhoa', 'Aspyn', 'Elina', 'Elsa', 'Magdalena', 'Kailey', 'Arleth', 'Joyce',
                 'Judith', 'Crystal', 'Emberlynn', 'Landry', 'Paola', 'Braylee', 'Guinevere', 'Aarna', 'Aiyana',
                 'Kahlani', 'Lyanna', 'Sariah', 'Itzayana', 'Aniya', 'Frida', 'Jaylene', 'Kiera', 'Loyalty', 'Azaria',
                 'Jaylee', 'Kamilah', 'Keyla', 'Kyleigh', 'Micah', 'Nataly', 'Kathleen', 'Zoya', 'Meghan', 'Soraya',
                 'Zoie', 'Arlette', 'Zola', 'Luisa', 'Vida', 'Ryder', 'Tatiana', 'Tori', 'Aarya', 'Eleanora', 'Sandra',
                 'Soleil', 'Annabella']

        #
        # names_url = "https://www.verywellfamily.com/top-1000-baby-girl-names-2757832"
        # data = requests.get(names_url)
        #
        # html_data = soup(data.text, "html.parser")
        # _namesData = html_data.body.find("ol", {"id": "mntl-sc-block_1-0-16"}).findAll("li")
        # for ele in _namesData:
        #     names.append(ele.getText())
        return names
